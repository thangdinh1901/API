"""Deep port audit: spec vs pcat vs xlsx for WN/SO/LJ flanges."""
from __future__ import annotations

import sqlite3
from pathlib import Path

PSPEC = Path(r"D:\04. Projects\06. NUI\NUI\Spec Sheets\NPMC.pspc")
PCAT = Path(r"D:\04. Projects\06. NUI\NUI\Spec Sheets\CATA_NUI.pcat")
ASME = Path(r"C:\AutoCAD Plant 3D 2026 Content\CPak ASME\ASME Pipes and Fittings Catalog.pcat")
XLSX = Path(r"D:\04. Projects\06. NUI\NUI\Spec Sheets\CATA_NUI.xlsx")


def ports_via_partport(conn: sqlite3.Connection, label: str, desc: str, nd: float) -> None:
    print(f"\n--- {label} {desc} ND={nd} (PartPort) ---")
    for row in conn.execute(
        """
        SELECT e.ShortDescription, pp.Name, pt.EndType, pt.PressureClass, pt.Facing, pt.NominalDiameter
        FROM EngineeringItems e
        JOIN PartPort pp ON pp.Part = e.PnPID
        JOIN Port pt ON pt.PnPID = pp.Port
        WHERE e.ShortDescription = ? AND ABS(e.NominalDiameter - ?) < 0.01
        ORDER BY pp.Name
        """,
        (desc, nd),
    ):
        print(" ", row)


def spec_ei_rows(conn: sqlite3.Connection, desc: str, nd: float) -> None:
    print(f"\n--- SPEC EI rows {desc} ND={nd} ---")
    for row in conn.execute(
        """
        SELECT e.PnPID, e.PortName, e.EndType, e.PressureClass, e.Facing, e.SizeRecordId
        FROM EngineeringItems e
        JOIN Flange f ON f.PnPID = e.PnPID
        WHERE e.ShortDescription = ? AND ABS(e.NominalDiameter - ?) < 0.01
        """,
        (desc, nd),
    ):
        print(" ", row)


def main() -> None:
    for desc in ("FLANGE WN", "FLANGE SO", "FLANGE LJ"):
        c = sqlite3.connect(PCAT)
        ports_via_partport(c, "CATA", desc, 50)
        c.close()
        c = sqlite3.connect(ASME)
        ports_via_partport(c, "ASME", desc.replace("FLANGE WN", "FLANGE WN").replace("FLANGE LJ", "FLANGE LJ"), 50 if "LJ" not in desc else 50)
        c.close()
        c = sqlite3.connect(PSPEC)
        spec_ei_rows(c, desc, 50)
        c.close()

    # ASME uses inches - WN at 2"
    c = sqlite3.connect(ASME)
    ports_via_partport(c, "ASME", "FLANGE WN", 2.0)
    ports_via_partport(c, "ASME", "FLANGE LJ", 2.0)
    c.close()

    # xlsx DN50 WN
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn.startswith("WN_FLRF"):
                ws = wb[sn]
                hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    d = dict(zip(hdr, row))
                    if str(d.get("DN")) == "50":
                        for k in sorted(d):
                            if d[k] not in (None, "") and ("Port" in k or "EndType" in k or "Pressure" in k or "Facing" in k or k in ("DN", "PnPClassName", "ContentGeometryParamDefinition")):
                                print(f"  xlsx {k}: {d[k]}")
                        break
    except Exception as e:
        print("xlsx error", e)


if __name__ == "__main__":
    main()
