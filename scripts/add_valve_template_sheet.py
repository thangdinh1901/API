"""Add VALVE_FL_CL150 clone sheet to CatalogBuilderTemplate.xlsx."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TPL = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"


def main() -> None:
    backup = TPL.with_suffix(".xlsx.bak")
    if not backup.exists():
        shutil.copy2(TPL, backup)

    wb = openpyxl.load_workbook(TPL)
    sheet_name = "VALVE_FL_CL150,FL,150"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    headers = [
        ("Previews", "Preview"),
        ("Sizes", "Sizes"),
        ("ShapeName", "Shape Name"),
        ("ScriptPath", "Script Path"),
        ("DN", "DN"),
        ("L", "L"),
        ("D1", "D1"),
        ("ContentGeometryParamDefinition", "ContentGeometryParamDefinition"),
        ("SizeRecordId_S-ALL", "Size Record Id_S-ALL"),
        ("PortName_S-ALL", "Port Name_S-ALL"),
        ("NominalDiameter_S-ALL", "Nominal Diameter_S-ALL"),
        ("NominalUnit_S-ALL", "Nominal Unit_S-ALL"),
        ("MatchingPipeOd_S-ALL", "Matching Pipe OD_S-ALL"),
        ("EndType_S-ALL", "End Type_S-ALL"),
        ("Facing_S-ALL", "Facing_S-ALL"),
        ("PressureClass_S-ALL", "Pressure Class_S-ALL"),
        ("Schedule_S-ALL", "Schedule_S-ALL"),
        ("WallThickness_S-ALL", "Wall Thickness_S-ALL"),
        ("EngagementLength_S-ALL", "Engagement Length_S-ALL"),
        ("LengthUnit_S-ALL", "Port Unit_S-ALL"),
        ("ShortDescription", "Short Description"),
        ("CompatibleStandard", "Compatible Standard"),
        ("DesignStd", "Design Std"),
        ("PartFamilyLongDesc", "Long Description (Family)"),
        ("PartSizeLongDesc", "Long Description (Size)"),
        ("Material", "Material"),
        ("MaterialCode", "Material Code"),
        ("Weight", "Weight"),
        ("WeightUnit", "Weight Unit"),
        ("PartFamilyId", "PartFamilyId"),
        ("CatalogPartFamilyId", "CatalogPartFamilyId"),
        ("ConnectionPortCount", "ConnectionPortCount"),
        ("PartCategory", "PartCategory"),
        ("PnPClassName", "PnPClassName"),
        ("SKEY", "SKEY"),
        ("TYPE", "TYPE"),
        ("ContentIsoSymbolDefinition", "ContentIsoSymbolDefinition"),
    ]

    for col, (machine, display) in enumerate(headers, start=1):
        ws.cell(1, col, machine)
        ws.cell(2, col, display)

    seed = {
        "ShapeName": "CUST_VALVE_FL_CL150",
        "ContentGeometryParamDefinition": "DN,L",
        "PortName_S-ALL": "ALL",
        "NominalUnit_S-ALL": "mm",
        "EndType_S-ALL": "FL",
        "Facing_S-ALL": "RF",
        "PressureClass_S-ALL": "150",
        "LengthUnit_S-ALL": "mm",
        "ShortDescription": "Custom valve",
        "CompatibleStandard": "ASME B16.10 / manufacturer",
        "DesignStd": "Custom",
        "PartFamilyLongDesc": "Valve. Flanged CL150 RF CS Custom (Plant 3D Composer template)",
        "Material": "CS",
        "ConnectionPortCount": "2",
        "PartCategory": "Valves",
        "PnPClassName": "Valve",
        "SKEY": "VFLG",
        "TYPE": "VALVE",
        "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VFLG",
    }
    name_to_col = {h[0]: i for i, h in enumerate(headers, start=1)}
    for key, val in seed.items():
        ws.cell(3, name_to_col[key], val)

    if "Catalog Data Flag" in wb.sheetnames:
        idx = wb.sheetnames.index("Catalog Data Flag")
        wb.move_sheet(ws, offset=-(len(wb.sheetnames) - idx))

    wb.save(TPL)
    print(f"Added sheet: {sheet_name} ({len(wb.sheetnames)} sheets total)")


if __name__ == "__main__":
    main()
