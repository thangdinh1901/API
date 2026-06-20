"""Ensure STUD_LJ catalog sheet exists (separate from STUD_RF for WN/SO)."""
from pathlib import Path

import openpyxl

TEMPLATE = Path(__file__).resolve().parents[1] / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"

wb = openpyxl.load_workbook(TEMPLATE)
src_name = next(s for s in wb.sheetnames if s.upper().startswith("STUD_RF"))
dst_name = "STUD_LJ_CL150,"

for old in list(wb.sheetnames):
    if old.upper().startswith("STUD_FF") or old.upper().startswith("STUD_LJ"):
        del wb[old]

src = wb[src_name]
dst = wb.copy_worksheet(src)
dst.title = dst_name
dst.cell(1, 1, dst_name)

wb.save(TEMPLATE)
print("Sheet", dst_name, "ready (cloned from", src_name + ")")
