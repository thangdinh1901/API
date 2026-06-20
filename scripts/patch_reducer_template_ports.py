"""Patch CatalogBuilderTemplate.xlsx: multi-port headers for reducers and tee reduce."""

from pathlib import Path

import openpyxl

TEMPLATE = Path(__file__).resolve().parents[1] / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"

PORT_FIELDS = (
    "SizeRecordId", "PortName", "NominalDiameter", "NominalUnit", "MatchingPipeOd",
    "EndType", "FlangeStd", "GasketStd", "Facing", "FlangeThickness",
    "PressureClass", "Schedule", "WallThickness", "EngagementLength", "LengthUnit",
)

SHEETS = {
    "REDUCER_CONC_BW_SCH40,FL,": 2,
    "REDUCER_ECC_BW_SCH40,FL,4": 2,
    "TEE_REDUCE_BW_SCH40,FL,40": 3,
    "TEE_REDUCE_SW_CL3000,FL,3": 3,
    "ELBOW_45_LR_BW_SCH40,FL,4": 2,
    "ELBOW_90_LR_BW_SCH40,FL,4": 2,
    "ELBOW_90_SR_BW_SCH40,FL,4": 2,
    "ELBOW_45_SW_CL3000,FL,300": 2,
    "ELBOW_90_SW_CL3000,FL,300": 2,
    "TEE_EQ_BW_SCH40,FL": 3,
    "TEE_EQ_SW_CL3000,FL,3000": 3,
}


def port_block_start(ws) -> int | None:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value and str(value).startswith("SizeRecordId_"):
            return col
    return None


def has_suffix(ws, suffix: str) -> bool:
    return any(ws.cell(1, c).value == f"SizeRecordId_{suffix}" for c in range(1, ws.max_column + 1))


def rename_block(ws, start: int, old: str, new: str) -> None:
    for offset, stem in enumerate(PORT_FIELDS):
        cell = ws.cell(1, start + offset)
        if cell.value and str(cell.value).endswith(f"_{old}"):
            cell.value = f"{stem}_{new}"


def insert_block(ws, start: int, suffix: str) -> None:
    ws.insert_cols(start, len(PORT_FIELDS))
    for offset, stem in enumerate(PORT_FIELDS):
        ws.cell(1, start + offset).value = f"{stem}_{suffix}"


def sync_row2(ws) -> None:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value and (
            str(value).startswith("SizeRecordId_")
            or any(str(value).startswith(f"{stem}_") for stem in PORT_FIELDS[1:])
        ):
            ws.cell(2, col).value = value


def patch_sheet(ws, port_count: int) -> None:
    start = port_block_start(ws)
    if start is None:
        raise RuntimeError(f"No port block in {ws.title}")

    first = str(ws.cell(1, start).value).split("_", 1)[1]
    if first == "S-ALL":
        rename_block(ws, start, "S-ALL", "S1")
        insert_block(ws, start + len(PORT_FIELDS), "S2")
        if port_count == 3:
            insert_block(ws, start + 2 * len(PORT_FIELDS), "S3")
    elif port_count == 2 and not has_suffix(ws, "S2"):
        insert_block(ws, start + len(PORT_FIELDS), "S2")
    elif port_count == 3:
        if not has_suffix(ws, "S2"):
            insert_block(ws, start + len(PORT_FIELDS), "S2")
        if not has_suffix(ws, "S3"):
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value == "SizeRecordId_S2":
                    insert_block(ws, col + len(PORT_FIELDS), "S3")
                    break

    sync_row2(ws)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    for name, ports in SHEETS.items():
        patch_sheet(wb[name], ports)
    wb.save(TEMPLATE)
    print(f"Patched {TEMPLATE}")


if __name__ == "__main__":
    main()
