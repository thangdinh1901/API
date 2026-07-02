"""One-off: fill PathAngle / CurveRadius / SegmentCount in an already-published
elbow-class catalog Excel so the native Plant 3D Elbow class has routing geometry.
Without these the catalog part has no symbol -> "Can't find symbol for specified part".
"""
import sys
import openpyxl

LR90_CENTER_TO_FACE_MM = {
    15: 38, 20: 38, 25: 38, 32: 48, 40: 57, 50: 76, 65: 95, 80: 114, 90: 133,
    100: 152, 125: 190, 150: 229, 200: 305, 250: 381, 300: 457, 350: 533,
    400: 610, 450: 686,
}

path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\dinht\Downloads\THANG_TEST.xlsx"
angle = float(sys.argv[2]) if len(sys.argv) > 2 else 90.0

wb = openpyxl.load_workbook(path)
patched_total = 0
for ws in wb.worksheets:
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            header[str(v).strip()] = c
    if "PathAngle" not in header or "CurveRadius" not in header or "DN" not in header:
        continue

    patched = 0
    for r in range(2, ws.max_row + 1):
        dn_val = ws.cell(r, header["DN"]).value
        if dn_val is None:
            continue
        try:
            dn = int(round(float(dn_val)))
        except (TypeError, ValueError):
            continue
        cr = LR90_CENTER_TO_FACE_MM.get(dn)
        if cr is None:
            cr = round(dn * 1.5, 1)
        ws.cell(r, header["PathAngle"]).value = angle
        ws.cell(r, header["CurveRadius"]).value = cr
        if "SegmentCount" in header:
            ws.cell(r, header["SegmentCount"]).value = 0
        patched += 1
    print(f"  {ws.title}: patched {patched} row(s)")
    patched_total += patched

out = sys.argv[3] if len(sys.argv) > 3 else path
try:
    wb.save(out)
except PermissionError:
    out = path.rsplit(".", 1)[0] + "_FIXED.xlsx"
    wb.save(out)
print(f"Saved {out} ({patched_total} rows)")

