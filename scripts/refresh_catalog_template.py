"""Refresh CatalogBuilderTemplate.xlsx — fix script paths and stud-bolt columns."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"
CUSTOM_SCRIPTS = Path(
    r"C:\AutoCAD Plant 3D 2026 Content\CPak Common\CustomScripts"
)

# Mirrors CatalogFlangeBoltingCatalog.cs (ASME B16.5 CL150 RF).
BOLTING_BY_DN: dict[int, tuple[str, int, float]] = {
    15: ('1/2"', 4, 55),
    20: ('1/2"', 4, 65),
    25: ('1/2"', 4, 65),
    32: ('1/2"', 4, 70),
    40: ('1/2"', 4, 70),
    50: ('5/8"', 4, 85),
    65: ('5/8"', 4, 90),
    80: ('5/8"', 4, 90),
    90: ('5/8"', 8, 90),
    100: ('5/8"', 8, 90),
    125: ('5/8"', 8, 95),
    150: ('3/4"', 8, 100),
    200: ('3/4"', 8, 110),
    250: ('7/8"', 12, 115),
    300: ('7/8"', 12, 120),
    350: ('1"', 12, 135),
    400: ('1"', 16, 135),
    450: ('1-1/8"', 16, 145),
}


def header_map(ws, header_row: int = 2) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        for row in (1, header_row):
            val = ws.cell(row, col).value
            if val is None:
                continue
            key = str(val).strip()
            if key and key not in mapping:
                mapping[key] = col
    return mapping


def set_cell(ws, hdr: dict[str, int], row: int, name: str, value) -> None:
    col = hdr.get(name)
    if col:
        ws.cell(row, col, value=value)


def fix_script_paths(ws, hdr: dict[str, int], data_start: int = 3) -> int:
    shape_col = hdr.get("Shape Name") or hdr.get("ShapeName")
    path_col = hdr.get("Script Path") or hdr.get("ScriptPath")
    if not shape_col or not path_col:
        return 0

    fixed = 0
    for row in range(data_start, ws.max_row + 1):
        shape = ws.cell(row, shape_col).value
        if not shape or not str(shape).startswith("CUST_"):
            continue
        py = CUSTOM_SCRIPTS / f"{shape}.py"
        if py.is_file():
            ws.cell(row, path_col, value=str(py))
            fixed += 1
    return fixed


def fill_stud_sheet(ws) -> int:
    hdr = header_map(ws)
    family_id = "36894312-7f7b-4dce-9b48-e62ceba5c6f9"
    updated = 0
    for row in range(3, ws.max_row + 1):
        sizes = ws.cell(row, hdr.get("Sizes", 2)).value
        if sizes is None:
            continue
        try:
            dn = int(sizes)
        except (TypeError, ValueError):
            continue
        bolting = BOLTING_BY_DN.get(dn)
        if not bolting:
            continue
        bolt_size, count, length_mm = bolting
        length_in = round(length_mm / 25.4, 6)
        set_cell(ws, hdr, row, "ShortDescription", "Bolt set")
        set_cell(ws, hdr, row, "PartFamilyLongDesc", "Studbolt ASTM A193-B7-Nuts ASTM A194-2H")
        set_cell(ws, hdr, row, "PartSizeLongDesc", f"Studbolt DN{dn} ASTM A193-B7-Nuts ASTM A194-2H")
        set_cell(ws, hdr, row, "PnPClassName", "BoltSet")
        set_cell(ws, hdr, row, "PartCategory", "Fasteners")
        set_cell(ws, hdr, row, "ConnectionPortCount", "1")
        set_cell(ws, hdr, row, "PartFamilyId", family_id)
        set_cell(ws, hdr, row, "CatalogPartFamilyId", family_id)
        set_cell(ws, hdr, row, "PressureClass_S-ALL", "150")
        set_cell(ws, hdr, row, "Facing_S-ALL", "RF")
        set_cell(ws, hdr, row, "Content Iso Symbol Definition", "TYPE=BOLT")
        set_cell(ws, hdr, row, "ContentIsoSymbolDefinition", "TYPE=BOLT")
        set_cell(ws, hdr, row, "TYPE", "BOLT")
        set_cell(ws, hdr, row, "SKEY", "BOLT")
        set_cell(ws, hdr, row, "BoltSize", bolt_size)
        set_cell(ws, hdr, row, "NumberInSet", count)
        set_cell(ws, hdr, row, "Length", length_in)
        set_cell(ws, hdr, row, "IsLugSet", "0")
        set_cell(ws, hdr, row, "StudTypeDescription", "Stud Bolt")
        set_cell(ws, hdr, row, "StudDescription", "Lg, ASTM A193, B7")
        set_cell(ws, hdr, row, "BoltCompatibleStd", "ASTM A193")
        set_cell(ws, hdr, row, "Material", "CS")
        updated += 1
    return updated


def set_flange_facing(ws) -> None:
    """S1/FL ports = RF (matches working pre-valve catalog). BV/SO ports = blank."""
    hdr = header_map(ws)
    for row in range(3, ws.max_row + 1):
        if not ws.cell(row, hdr.get("Sizes", 2)).value:
            continue
        for name, col in hdr.items():
            if not name or "Facing" not in str(name):
                continue
            suffix = str(name).replace("Facing_", "").replace("Facing ", "")
            end_col = hdr.get(f"EndType_{suffix}") or hdr.get(f"End Type_{suffix}")
            end_type = ws.cell(row, end_col).value if end_col else None
            if end_type and str(end_type).upper() == "FL":
                ws.cell(row, col, value="RF")
            else:
                ws.cell(row, col, value=None)


def main() -> None:
    if not TEMPLATE.is_file():
        raise SystemExit(f"Template not found: {TEMPLATE}")

    backup = TEMPLATE.with_suffix(".xlsx.bak")
    shutil.copy2(TEMPLATE, backup)

    wb = openpyxl.load_workbook(TEMPLATE)
    script_fixes = 0
    for name in wb.sheetnames:
        ws = wb[name]
        hdr = header_map(ws)
        script_fixes += fix_script_paths(ws, hdr)
        upper = name.upper()
        if upper.startswith("STUD_RF"):
            fill_stud_sheet(ws)
        elif any(upper.startswith(p) for p in ("WN_", "SO_", "BLD_", "ELBOW_", "TEE_", "REDUCER_")):
            set_flange_facing(ws)
        elif upper.startswith("GSK_"):
            set_flange_facing(ws)

    wb.save(TEMPLATE)
    wb.close()
    print(f"Backed up -> {backup}")
    print(f"Fixed {script_fixes} script path(s); refreshed STUD_RF bolting columns.")
    print(f"Saved {TEMPLATE}")


if __name__ == "__main__":
    main()
