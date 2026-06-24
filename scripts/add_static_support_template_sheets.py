"""Restore native Plant 3D support sheets (pipe CPP, stud) in CatalogBuilderTemplate.xlsx.

Pipe stock is NOT a custom CUST_*.py part — Catalog Builder expects ShapeName=CPP
(built-in Cylinder Pipe Parametric) with ContentGeometryParamDefinition=D,L,OF,
"""
from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
TPL = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"
GIT_ROOT = ROOT
NATIVE_PIPE_SHEET = "PIPE_SCH40,BV,40"
LEGACY_BAD_PIPE = "PIPE_SCH40,PL,150"
NATIVE_SOURCE_REV = "365f732"


def _extract_template(rev: str, dest: Path) -> None:
    proc = subprocess.run(
        ["git", "-C", str(GIT_ROOT), "show", f"{rev}:{TPL.relative_to(GIT_ROOT).as_posix()}"],
        capture_output=True,
        check=True,
    )
    dest.write_bytes(proc.stdout)


def _copy_sheet(src: Worksheet, dest_wb: openpyxl.Workbook, name: str) -> None:
    if name in dest_wb.sheetnames:
        del dest_wb[name]
    dest = dest_wb.create_sheet(name)
    for row in src.iter_rows():
        for cell in row:
            dest.cell(cell.row, cell.column, value=cell.value)
    for col, dim in src.column_dimensions.items():
        dest.column_dimensions[col].width = dim.width
        dest.column_dimensions[col].hidden = dim.hidden
    for row_idx, dim in src.row_dimensions.items():
        dest.row_dimensions[row_idx].hidden = dim.hidden


def unhide_header_rows(wb: openpyxl.Workbook) -> int:
    changed = 0
    for ws in wb.worksheets:
        rd = ws.row_dimensions.get(1)
        if rd and rd.hidden:
            rd.hidden = False
            changed += 1
    return changed


def restore_native_support_sheets(wb: openpyxl.Workbook) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        src_path = Path(tmp) / "native_template.xlsx"
        _extract_template(NATIVE_SOURCE_REV, src_path)
        src_wb = openpyxl.load_workbook(src_path)

        for sheet_name in (NATIVE_PIPE_SHEET, "STUD_RF_CL150,", "STUD_LJ_CL150,"):
            if sheet_name not in src_wb.sheetnames:
                print(f"WARN: missing native sheet {sheet_name} in {NATIVE_SOURCE_REV}")
                continue
            _copy_sheet(src_wb[sheet_name], wb, sheet_name)
            print(f"Restored {sheet_name}")

        if LEGACY_BAD_PIPE in wb.sheetnames:
            del wb[LEGACY_BAD_PIPE]
            print(f"Removed {LEGACY_BAD_PIPE}")

        src_wb.close()

    if "Catalog Data Flag" in wb.sheetnames:
        flag_idx = wb.sheetnames.index("Catalog Data Flag")
        for name in (NATIVE_PIPE_SHEET, "STUD_RF_CL150,", "STUD_LJ_CL150,"):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=-(len(wb.sheetnames) - flag_idx))


def main() -> int:
    if not TPL.is_file():
        raise SystemExit(f"Template not found: {TPL}")
    backup = TPL.with_suffix(".xlsx.bak")
    if not backup.exists():
        shutil.copy2(TPL, backup)

    wb = openpyxl.load_workbook(TPL)
    restore_native_support_sheets(wb)
    unhidden = unhide_header_rows(wb)
    wb.save(TPL)
    print(f"Unhid row 1 on {unhidden} sheet(s)")
    print(f"Saved {TPL}")

    valve_script = ROOT / "scripts" / "add_valve_template_sheets.py"
    if valve_script.is_file():
        import sys
        subprocess.run([sys.executable, str(valve_script)], check=True, cwd=str(ROOT))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
