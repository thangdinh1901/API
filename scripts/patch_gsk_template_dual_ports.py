"""Patch CatalogBuilderTemplate.xlsx — GSK sheets dual FL ports (row 1 + row 2 must match)."""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"
WN_SHEET = "WN_FLRF_CL150,FL,150"
GSK_SHEETS = ("GSK_RF_CL150,FL,150", "GSK_FF_CL150,FL,150")

GSK_S1_START = 8
GSK_S2_START = 23
WN_S1_START = 7
WN_S1_LEN = 15
WN_S2_START = 22
WN_S2_LEN = 15
META_WN_START = 37  # ShortDescription...


def _copy_row1_meta(gsk, wn, gsk_meta_start: int) -> None:
    wn_meta = [wn.cell(1, c).value for c in range(META_WN_START, wn.max_column + 1)]
    for offset, val in enumerate(wn_meta):
        gsk.cell(1, gsk_meta_start + offset, value=val)


def ensure_gsk_dual_ports(wb: openpyxl.Workbook, gsk_name: str) -> None:
    if gsk_name not in wb.sheetnames:
        print(f"WARN: missing sheet {gsk_name}")
        return
    wn = wb[WN_SHEET]
    gsk = wb[gsk_name]

    hdr2 = [gsk.cell(2, c).value for c in range(1, gsk.max_column + 1)]
    has_s2 = any(v and "EndType_S2" in str(v) for v in hdr2)

    if not has_s2:
        for offset in range(WN_S1_LEN):
            val = wn.cell(2, WN_S1_START + offset).value
            gsk.cell(2, GSK_S1_START + offset, value=val)
        gsk.insert_cols(GSK_S2_START, WN_S2_LEN)
        for offset in range(WN_S2_LEN):
            gsk.cell(2, GSK_S2_START + offset, value=wn.cell(2, WN_S2_START + offset).value)

    # Row 2: ensure S1/S2 names (not S-ALL).
    for col in range(GSK_S1_START, GSK_S2_START + WN_S2_LEN):
        for row in (1, 2):
            val = gsk.cell(row, col).value
            if val and "_S-ALL" in str(val):
                gsk.cell(row, col, value=str(val).replace("_S-ALL", "_S1"))

    # Row 1 S1 block from WN S1 (offset +1 for GSK col T).
    for offset in range(WN_S1_LEN):
        gsk.cell(1, GSK_S1_START + offset, value=wn.cell(1, WN_S1_START + offset).value)

    # Row 1 S2 block from WN S2.
    for offset in range(WN_S2_LEN):
        gsk.cell(1, GSK_S2_START + offset, value=wn.cell(1, WN_S2_START + offset).value)

    gsk_meta_start = GSK_S2_START + WN_S2_LEN  # 38
    _copy_row1_meta(gsk, wn, gsk_meta_start)

    # Row 2 metadata labels already shifted by insert; refresh from row 1 where blank.
    for col in range(gsk_meta_start, gsk.max_column + 1):
        if not gsk.cell(2, col).value and wn.cell(2, col - 1).value:
            gsk.cell(2, col, value=wn.cell(2, col - 1).value)

    print(f"Patched {gsk_name} (row 1 + row 2 dual ports)")


def main() -> int:
    if not TEMPLATE.is_file():
        raise SystemExit(f"Template not found: {TEMPLATE}")
    wb = openpyxl.load_workbook(TEMPLATE)
    for name in GSK_SHEETS:
        ensure_gsk_dual_ports(wb, name)
    wb.save(TEMPLATE)
    print(f"Saved {TEMPLATE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
