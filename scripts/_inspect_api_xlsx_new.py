import openpyxl
from openpyxl.utils import get_column_letter

path = r"c:\Users\dinht\Downloads\API.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} === rows={ws.max_row} cols={ws.max_column} state={ws.sheet_state}")
    for r in range(1, min(4, ws.max_row + 1)):
        cols = []
        for c in range(1, min(ws.max_column + 1, 100)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                cols.append(f"{get_column_letter(c)}={v!r}")
        if cols:
            print(f"  row{r}: {', '.join(cols[:35])}")
            if len(cols) > 35:
                print(f"    ... +{len(cols)-35} more")
    if ws.max_row >= 3:
        print("  sample row3:")
        cols = []
        for c in range(1, min(ws.max_column + 1, 100)):
            v = ws.cell(3, c).value
            if v is not None and str(v).strip():
                cols.append(f"{get_column_letter(c)}={v!r}")
        if cols:
            print(f"    {', '.join(cols[:40])}")
wb.close()
