"""Add COLLAR_LJ_* sheets to CatalogBuilderTemplate (clone from STUBEND_LJ_*)."""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"

PAIRS = (
    ("STUBEND_LJ_A_BW_SCH40,FL,40", "COLLAR_LJ_A_BW_SCH40,FL,40"),
    ("STUBEND_LJ_A_SH_BW_SCH40,FL,40", "COLLAR_LJ_A_SH_BW_SCH40,FL,40"),
)


def _copy_sheet(wb: openpyxl.Workbook, src_name: str, dst_name: str) -> None:
    if dst_name in wb.sheetnames:
        print(f"  exists: {dst_name}")
        return
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    dst.title = dst_name
    # Row 1 shape / template hints
    for row in (1, 2, 3):
        for col in range(1, src.max_column + 1):
            val = src.cell(row, col).value
            if val and isinstance(val, str):
                dst.cell(row, col, value=val.replace("STUBEND", "COLLAR"))
    print(f"  added: {dst_name}")


def main() -> int:
    if not TEMPLATE.is_file():
        raise SystemExit(f"Template not found: {TEMPLATE}")
    wb = openpyxl.load_workbook(TEMPLATE)
    for src, dst in PAIRS:
        if src not in wb.sheetnames:
            raise SystemExit(f"Missing source sheet: {src}")
        _copy_sheet(wb, src, dst)
    wb.save(TEMPLATE)
    print(f"Saved {TEMPLATE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
