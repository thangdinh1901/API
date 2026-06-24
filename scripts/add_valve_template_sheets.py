"""Add Plant 3D Composer valve clone sheets to CatalogBuilderTemplate.xlsx."""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TPL = ROOT / "Plant3DCatalogComposer" / "Resources" / "CatalogBuilderTemplate.xlsx"

META_HEADERS = [
    ("ShortDescription", "Short Description"),
    ("CompatibleStandard", "Compatible Standard"),
    ("DesignStd", "Design Std"),
    ("PartFamilyLongDesc", "Long Description (Family)"),
    ("PartSizeLongDesc", "Long Description (Size)"),
    ("Material", "Material"),
    ("MaterialCode", "Material Code"),
    ("Weight", "Weight"),
    ("WeightUnit", "Weight Unit"),
    ("PartFamilyId", "PartFamilyId"),
    ("CatalogPartFamilyId", "CatalogPartFamilyId"),
    ("ConnectionPortCount", "ConnectionPortCount"),
    ("PartCategory", "PartCategory"),
    ("PnPClassName", "PnPClassName"),
    ("SKEY", "SKEY"),
    ("TYPE", "TYPE"),
    ("ContentIsoSymbolDefinition", "ContentIsoSymbolDefinition"),
]

PORT_S_ALL = [
    ("SizeRecordId_S-ALL", "Size Record Id_S-ALL"),
    ("PortName_S-ALL", "Port Name_S-ALL"),
    ("NominalDiameter_S-ALL", "Nominal Diameter_S-ALL"),
    ("NominalUnit_S-ALL", "Nominal Unit_S-ALL"),
    ("MatchingPipeOd_S-ALL", "Matching Pipe OD_S-ALL"),
    ("EndType_S-ALL", "End Type_S-ALL"),
    ("Facing_S-ALL", "Facing_S-ALL"),
    ("PressureClass_S-ALL", "Pressure Class_S-ALL"),
    ("Schedule_S-ALL", "Schedule_S-ALL"),
    ("WallThickness_S-ALL", "Wall Thickness_S-ALL"),
    ("EngagementLength_S-ALL", "Engagement Length_S-ALL"),
    ("LengthUnit_S-ALL", "Port Unit_S-ALL"),
]

PORT_S1_S2 = [
    ("SizeRecordId_S1", "Size Record Id_S1"),
    ("PortName_S1", "Port Name_S1"),
    ("NominalDiameter_S1", "Nominal Diameter_S1"),
    ("NominalUnit_S1", "Nominal Unit_S1"),
    ("MatchingPipeOd_S1", "Matching Pipe OD_S1"),
    ("EndType_S1", "End Type_S1"),
    ("Facing_S1", "Facing_S1"),
    ("PressureClass_S1", "Pressure Class_S1"),
    ("Schedule_S1", "Schedule_S1"),
    ("WallThickness_S1", "Wall Thickness_S1"),
    ("EngagementLength_S1", "Engagement Length_S1"),
    ("LengthUnit_S1", "Port Unit_S1"),
    ("SizeRecordId_S2", "Size Record Id_S2"),
    ("PortName_S2", "Port Name_S2"),
    ("NominalDiameter_S2", "Nominal Diameter_S2"),
    ("NominalUnit_S2", "Nominal Unit_S2"),
    ("MatchingPipeOd_S2", "Matching Pipe OD_S2"),
    ("EndType_S2", "End Type_S2"),
    ("Facing_S2", "Facing_S2"),
    ("PressureClass_S2", "Pressure Class_S2"),
    ("Schedule_S2", "Schedule_S2"),
    ("WallThickness_S2", "Wall Thickness_S2"),
    ("EngagementLength_S2", "Engagement Length_S2"),
    ("LengthUnit_S2", "Port Unit_S2"),
]


def build_headers(geo: list[tuple[str, str]], ports: list[tuple[str, str]]) -> list[tuple[str, str]]:
    base = [
        ("Previews", "Preview"),
        ("Sizes", "Sizes"),
        ("ShapeName", "Shape Name"),
        ("ScriptPath", "Script Path"),
        ("DN", "DN"),
    ]
    return base + geo + [("ContentGeometryParamDefinition", "ContentGeometryParamDefinition")] + ports + META_HEADERS


def write_sheet(wb: openpyxl.Workbook, spec: dict[str, Any]) -> None:
    name = spec["sheet"]
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    headers = build_headers(spec["geo"], spec["ports"])
    for col, (machine, display) in enumerate(headers, start=1):
        ws.cell(1, col, machine)
        ws.cell(2, col, display)
    col_map = {h[0]: i for i, h in enumerate(headers, start=1)}
    seed = dict(spec["seed"])
    seed.setdefault("ShapeName", f"CUST_{spec['part_id']}")
    seed.setdefault("ContentGeometryParamDefinition", spec["cgp"])
    seed.setdefault("PartCategory", "Valves")
    seed.setdefault("PnPClassName", "Valve")
    seed.setdefault("Material", "CS")
    seed.setdefault("NominalUnit_S-ALL", "mm")
    seed.setdefault("LengthUnit_S-ALL", "mm")
    seed.setdefault("NominalUnit_S1", "mm")
    seed.setdefault("NominalUnit_S2", "mm")
    seed.setdefault("LengthUnit_S1", "mm")
    seed.setdefault("LengthUnit_S2", "mm")
    for key, val in seed.items():
        if key in col_map:
            ws.cell(3, col_map[key], val)


VALVE_TEMPLATES: list[dict[str, Any]] = [
    {
        "part_id": "VALVE_FL_CL150",
        "sheet": "VALVE_FL_CL150,FL,150",
        "geo": [("L", "L"), ("D1", "D1")],
        "cgp": "DN,L",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "Custom flanged valve",
            "CompatibleStandard": "ASME B16.10 / manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Flanged CL150 RF CS Custom — minimal (DN + face-to-face)",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "FL",
            "Facing_S-ALL": "RF",
            "PressureClass_S-ALL": "150",
            "ConnectionPortCount": "2",
            "SKEY": "VFLG",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VFLG",
        },
    },
    {
        "part_id": "VALVE_FL_RICH",
        "sheet": "VALVE_FL_RICH,FL,150",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("L", "L"),
            ("LS", "LS"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("OF", "OF"),
            ("B1", "B1"),
            ("B2", "B2"),
        ],
        "cgp": "D1,D2,L,LS,H1,H2,W1,W2,OF,B1,B2,",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "Flanged valve (rich dims)",
            "CompatibleStandard": "ASME B16.10 / manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Flanged CL150 RF CS — gate/globe/plug style (Plant-like geometry columns)",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "FL",
            "Facing_S-ALL": "RF",
            "PressureClass_S-ALL": "150",
            "ConnectionPortCount": "2",
            "SKEY": "VFLG",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VFLG",
        },
    },
    {
        "part_id": "VALVE_BV_CL150",
        "sheet": "VALVE_BV_CL150,BV,150",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("L", "L"),
            ("LS", "LS"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("D3", "D3"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("OF", "OF"),
            ("B1", "B1"),
            ("B2", "B2"),
        ],
        "cgp": "D1,D2,L,LS,H1,H2,D3,W1,W2,OF,B1,B2,",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "Butt-weld valve",
            "CompatibleStandard": "ASME B16.10 / manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Butt-weld CL150 CS — ball/plug BW ends",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "BV",
            "PressureClass_S-ALL": "150",
            "Schedule_S-ALL": "40",
            "ConnectionPortCount": "2",
            "SKEY": "VBWG",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VBWG",
        },
    },
    {
        "part_id": "VALVE_SW_CL3000",
        "sheet": "VALVE_SW_CL3000,SW,3000",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("L", "L"),
            ("LS", "LS"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("D3", "D3"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("OF", "OF"),
            ("L1", "L1"),
            ("L2", "L2"),
            ("I1", "I1"),
            ("I2", "I2"),
            ("B1", "B1"),
            ("B2", "B2"),
        ],
        "cgp": "D1,D2,L,LS,H1,H2,D3,W1,W2,OF,L1,L2,I1,I2,B1,B2,",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "Socket-weld valve",
            "CompatibleStandard": "ASME B16.11 / manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Socket-weld CL3000 CS",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "SW",
            "PressureClass_S-ALL": "3000",
            "ConnectionPortCount": "2",
            "SKEY": "VSWG",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VSWG",
        },
    },
    {
        "part_id": "VALVE_3WAY",
        "sheet": "VALVE_3WAY,FL,150",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("D3", "D3"),
            ("L1", "L1"),
            ("L2", "L2"),
            ("L3", "L3"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("H3", "H3"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("W3", "W3"),
            ("A", "A"),
            ("OF", "OF"),
            ("B1", "B1"),
            ("B2", "B2"),
            ("B3", "B3"),
            ("P", "P"),
        ],
        "cgp": "D1,D2,D3,L1,L2,L3,H1,H2,H3,W1,W2,W3,A,OF,B1,B2,B3,P,",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "3-way valve",
            "CompatibleStandard": "Manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. 3-way multi-port CS",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "FL",
            "Facing_S-ALL": "RF",
            "PressureClass_S-ALL": "150",
            "ConnectionPortCount": "3",
            "SKEY": "V3WY",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=V3WY",
        },
    },
    {
        "part_id": "VALVE_ANGLE",
        "sheet": "VALVE_ANGLE,BV,150",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("L1", "L1"),
            ("L2", "L2"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("A", "A"),
            ("OF", "OF"),
            ("B1", "B1"),
            ("B2", "B2"),
        ],
        "cgp": "D1,D2,L1,L2,H1,H2,W1,W2,A,OF,B1,B2,",
        "ports": PORT_S_ALL,
        "seed": {
            "ShortDescription": "Angle valve",
            "CompatibleStandard": "Manufacturer",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Angle pattern CS",
            "PortName_S-ALL": "ALL",
            "EndType_S-ALL": "BV",
            "PressureClass_S-ALL": "150",
            "Schedule_S-ALL": "40",
            "ConnectionPortCount": "2",
            "SKEY": "VANG",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VANG",
        },
    },
    {
        "part_id": "VALVE_PSV",
        "sheet": "VALVE_PSV,FL,150",
        "geo": [
            ("D1", "D1"),
            ("D2", "D2"),
            ("L1", "L1"),
            ("L2", "L2"),
            ("X1", "X1"),
            ("X2", "X2"),
            ("H1", "H1"),
            ("H2", "H2"),
            ("W1", "W1"),
            ("W2", "W2"),
            ("A", "A"),
            ("OF", "OF"),
            ("B1", "B1"),
            ("B2", "B2"),
            ("DA", "DA"),
            ("HA", "HA"),
            ("isPilot", "isPilot"),
        ],
        "cgp": "D1,D2,L1,L2,X1,X2,H1,H2,W1,W2,A,OF,B1,B2,DA,HA,isPilot,",
        "ports": PORT_S1_S2,
        "seed": {
            "ShortDescription": "Pressure relief valve",
            "CompatibleStandard": "ASME VIII / API 526",
            "DesignStd": "Custom",
            "PartFamilyLongDesc": "Valve. Pressure relief / safety CS",
            "PortName_S1": "S1",
            "PortName_S2": "S2",
            "EndType_S1": "FL",
            "EndType_S2": "FL",
            "Facing_S1": "RF",
            "Facing_S2": "RF",
            "PressureClass_S1": "150",
            "PressureClass_S2": "150",
            "ConnectionPortCount": "2",
            "SKEY": "VPSV",
            "TYPE": "VALVE",
            "ContentIsoSymbolDefinition": "TYPE=VALVE,SKEY=VPSV",
        },
    },
]


def main() -> None:
    backup = TPL.with_suffix(".xlsx.bak")
    if not backup.exists():
        shutil.copy2(TPL, backup)

    wb = openpyxl.load_workbook(TPL)
    for spec in VALVE_TEMPLATES:
        write_sheet(wb, spec)
        print(f"  + {spec['sheet']}")

    if "Catalog Data Flag" in wb.sheetnames:
        flag_idx = wb.sheetnames.index("Catalog Data Flag")
        valve_sheets = [spec["sheet"] for spec in VALVE_TEMPLATES]
        for sheet_name in valve_sheets:
            if sheet_name in wb.sheetnames:
                wb.move_sheet(sheet_name, offset=-(len(wb.sheetnames) - flag_idx))

    wb.save(TPL)
    print(f"Saved {TPL.name} ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
