#!/usr/bin/env python3
"""Generate NATIVE catalog parts from CATA_NUI.xlsx (Đợt 1: SIMPLE shapes).

For each supported sheet: read the size rows + ShapeName, build a static
DIMENSIONS table, and emit a part folder under catalog_generator/parts/<id>/:
    part.json, catalog_entry.py, <id>/CUST_<id>.py

Geometry is delegated to native_shapes.Native* (deployed as a shared module).
Run: python scripts/gen_native_parts.py  [--excel PATH] [--dry]
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
GEN = os.path.normpath(os.path.join(HERE, "..", "catalog_generator"))
PARTS = os.path.join(GEN, "parts")
DEFAULT_XLSX = r"D:/04. Projects/06. NUI/NUI/Spec Sheets/CATA_NUI.xlsx"

sys.path.insert(0, GEN)
import pipe_sizes  # noqa: E402


# sheet ShapeName -> (native class, geometry-arg builder). Builder returns the
# ordered ctor args (after s) as a list of (pyname, value) for the DIMENSIONS row.
SIMPLE_SHAPES = {
    "CPFWR_F_SF": "NativeSlipOnFlange",
    "CPFWR": "NativeWeldNeckFlange",
    "CPFLR": "NativeLapJointFlange",
    "CPFBR": "NativeBlindFlange",
    "CPP": "NativePipe",
    "CPJRC": "NativeReducerConc",
    "CPJRE": "NativeReducerEcc",
    "CPMUW": "NativeStubEnd",
    "CPB": "NativeElbow",
    "CPB_OFOF": "NativeElbowSocket",
    "CPTS": "NativeTee",
    "CPTS_OFOFOF": "NativeTeeSocket",
}


def _col_index(rows):
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    return {h: i for i, h in enumerate(hdr)}


def _num(v, default=0.0):
    if v is None or str(v).strip() == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _wall(dn):
    """Nominal wall (mm) from default spec schedule; 0 if unknown."""
    try:
        return float(pipe_sizes.nominal_wall_mm_default_spec(int(round(dn))))
    except Exception:
        return 0.0


def _pipe_od(r, ci, dn):
    """Matching pipe OD (mm) for SW body diameter; falls back to SCH40 OD by DN."""
    for col in ("MatchingPipeOd_S-ALL", "MatchingPipeOd_S1"):
        if col in ci and r[ci[col]]:
            try:
                return float(r[ci[col]])
            except (TypeError, ValueError):
                pass
    try:
        return float(pipe_sizes.pipe_od_sch40_mm(int(round(dn))))
    except Exception:
        return 0.0


def _size_rows(rows, ci):
    """Yield data rows that have a ShapeName."""
    si = ci.get("ShapeName")
    for r in rows[1:]:
        if r and si is not None and r[si] and str(r[si]).strip() not in ("ShapeName", "Shape Name"):
            yield r


def _sanitize(name):
    return re.sub(r"[^A-Za-z0-9_]", "_", name.strip()).strip("_")


def _row_dims(shape, r, ci):
    """Return (dn, dn2_or_None, args_tuple) for a size row, args ordered per class ctor."""
    g = lambda k: _num(r[ci[k]]) if k in ci else 0.0
    # Single-port sheets use _S1; blind/pipe use _S-ALL. Fall back across both.
    dn_col = "NominalDiameter_S1" if "NominalDiameter_S1" in ci else (
        "NominalDiameter_S-ALL" if "NominalDiameter_S-ALL" in ci else None)
    dn = int(round(_num(r[ci[dn_col]]))) if dn_col else 0
    dn2 = None

    if shape in ("CPFWR_F_SF", "CPFWR"):
        # (L, B, D1, D2, D3, socket_depth)
        I = g("I") if "I" in ci else 0.0
        return dn, None, (g("L"), g("B"), g("D1"), g("D2"), g("D3"), I)
    if shape == "CPFLR":
        return dn, None, (g("L"), g("D1"), g("D2"))
    if shape == "CPFBR":
        return dn, None, (g("L"), g("D"))
    if shape == "CPP":
        d = g("D")
        di = d - 2.0 * _wall(dn)
        L = g("L")
        if L <= 0:
            L = d * 3.0
        return dn, None, (d, round(L, 3), round(di, 3))
    if shape in ("CPJRC", "CPJRE"):
        dn2 = int(round(_num(r[ci["NominalDiameter_S2"]]))) if "NominalDiameter_S2" in ci else 0
        d1, d2 = g("D1"), g("D2")
        w1, w2 = _wall(dn), _wall(dn2 or dn)
        if shape == "CPJRE":
            return dn, dn2, (d1, d2, g("L"), g("E"), round(w1, 3), round(w2, 3))
        return dn, dn2, (d1, d2, g("L"), round(w1, 3), round(w2, 3))
    if shape == "CPMUW":
        # Stub-end barrel wall = lap thickness B (pipe wall of the flared collar),
        # per ASME B16.9 / stubend_geom.py — not the default-spec schedule wall.
        return dn, None, (g("L"), g("B"), g("D1"), g("D2"), round(g("B"), 3))
    if shape == "CPB":
        # BW elbow: D, R(bend), A, wall (from schedule). L1=L2=0.
        return dn, None, (g("D"), g("R"), g("A") or 90.0, round(_wall(dn), 3))
    if shape == "CPB_OFOF":
        # SW elbow: D(forging OD), R, A, L1, I1, L2, I2, wall, socket_bore, body_od(pipe).
        d = g("D")
        pod = _pipe_od(r, ci, dn)
        try:
            sb = float(pipe_sizes.sw_cl3000_socket_bore_mm(dn))
        except Exception:
            sb = g("D2") or 0.0
        w = round(pipe_sizes.nominal_wall_mm_default_spec(dn), 3) if pod else round(_wall(dn), 3)
        return dn, None, (d, g("R"), g("A") or 90.0, g("L1"), g("I1"),
                          g("L2"), g("I2"), w, round(sb, 3), round(pod, 3))
    if shape == "CPTS":
        # BW tee: D1(run), D3(branch), L1, L2, L3, A, wall. Reducing → branch DN2.
        dn2 = None
        if "NominalDiameter_S3" in ci and r[ci["NominalDiameter_S3"]]:
            dn2 = int(round(_num(r[ci["NominalDiameter_S3"]])))
        return dn, dn2, (g("D1"), g("D3"), g("L1"), g("L2"), g("L3"),
                         g("A") or 90.0, round(_wall(dn), 3))
    if shape == "CPTS_OFOFOF":
        # SW tee: D1, D3, L1, L2, L3, I1, I2, I3, A, wall, socket_bore, body_od, body_od3.
        dn2 = None
        if "NominalDiameter_S3" in ci and r[ci["NominalDiameter_S3"]]:
            dn2 = int(round(_num(r[ci["NominalDiameter_S3"]])))
        pod = _pipe_od(r, ci, dn)
        pod3 = pipe_sizes.pipe_od_sch40_mm(dn2) if dn2 else pod
        try:
            sb = float(pipe_sizes.sw_cl3000_socket_bore_mm(dn))
        except Exception:
            sb = 0.0
        w = round(pipe_sizes.nominal_wall_mm_default_spec(dn), 3) if pod else round(_wall(dn), 3)
        return dn, dn2, (g("D1"), g("D3"), g("L1"), g("L2"), g("L3"),
                         g("I1"), g("I2"), g("I3"), g("A") or 90.0,
                         w, round(sb, 3), round(pod, 3), round(pod3, 3))
    raise ValueError(f"Unsupported shape {shape}")


CLASS_CALL = {
    "NativeWeldNeckFlange": "L, B, D1, D2, D3, socket_depth=I",
    "NativeSlipOnFlange": "L, B, D1, D2, D3, socket_depth=I",
    "NativeLapJointFlange": "L, D1, D2",
    "NativeBlindFlange": "L, D",
    "NativePipe": "D, L, DI",
    "NativeReducerConc": "D1, D2, L, wall1, wall2",
    "NativeReducerEcc": "D1, D2, L, E, wall1, wall2",
    "NativeStubEnd": "L, B, D1, D2, wall",
    "NativeElbow": "D, R, A, wall",
    "NativeElbowSocket": "D, R, A, L1, I1, L2, I2, wall, socket_bore, body_od",
    "NativeTee": "D1, D3, L1, L2, L3, A, wall",
    "NativeTeeSocket": "D1, D3, L1, L2, L3, I1, I2, I3, A, wall, socket_bore, body_od, body_od3",
}
ARG_NAMES = {
    "NativeWeldNeckFlange": ["L", "B", "D1", "D2", "D3", "I"],
    "NativeSlipOnFlange": ["L", "B", "D1", "D2", "D3", "I"],
    "NativeLapJointFlange": ["L", "D1", "D2"],
    "NativeBlindFlange": ["L", "D"],
    "NativePipe": ["D", "L", "DI"],
    "NativeReducerConc": ["D1", "D2", "L", "wall1", "wall2"],
    "NativeReducerEcc": ["D1", "D2", "L", "E", "wall1", "wall2"],
    "NativeStubEnd": ["L", "B", "D1", "D2", "wall"],
    "NativeElbow": ["D", "R", "A", "wall"],
    "NativeElbowSocket": ["D", "R", "A", "L1", "I1", "L2", "I2", "wall", "socket_bore", "body_od"],
    "NativeTee": ["D1", "D3", "L1", "L2", "L3", "A", "wall"],
    "NativeTeeSocket": ["D1", "D3", "L1", "L2", "L3", "I1", "I2", "I3", "A", "wall", "socket_bore", "body_od", "body_od3"],
}


def _emit_part(part_id, cls, meta, table, reducer, dry):
    part_dir = os.path.join(PARTS, part_id)
    geom_dir = os.path.join(part_dir, part_id)
    argnames = ARG_NAMES[cls]
    call = CLASS_CALL[cls]
    py_class = _sanitize(part_id)

    # DIMENSIONS literal
    lines = []
    for key in sorted(table):
        lines.append(f"    {key!r}: {tuple(table[key])!r},")
    dims = "\n".join(lines)
    keydoc = "(DN, DN2)" if reducer else "DN"
    resolve = (
        "        key = (pipe_sizes.resolve_dn(size), pipe_sizes.resolve_dn(size2))\n"
        if reducer else
        "        key = pipe_sizes.resolve_dn(size)\n"
    )
    sig = "self, s, size, size2=None, *, add_ports=True" if reducer else "self, s, size, *, add_ports=True"
    unpack = ", ".join(argnames)

    geom = f'''"""{meta['long']}

NATIVE reconstruction of Plant 3D {meta['shape']}. Dimensions from CATA_NUI.xlsx
sheet "{meta['sheet']}". Auto-generated by scripts/gen_native_parts.py — do not edit by hand.
"""

import native_shapes
import pipe_sizes

# {keydoc} -> ({", ".join(argnames)})
DIMENSIONS = {{
{dims}
}}


class {py_class}(native_shapes.{cls}):
    def __init__({sig}):
{resolve}        if key not in DIMENSIONS:
            raise ValueError(f"No {part_id} data for {{key}}.")
        {unpack} = DIMENSIONS[key]
        super().__init__(s, {call}, add_ports=add_ports)
'''

    dn_param = f"DN={meta['defaultDN']}"
    if reducer:
        entry_args = "s, pipe_sizes.resolve_dn(DN), size2=DN2, add_ports=not preview"
        activate_sig = f"s, DN={meta['defaultDN']}, DN2={meta['defaultDN2']}, **kw"
    else:
        entry_args = "s, pipe_sizes.resolve_dn(DN), add_ports=not preview"
        activate_sig = f"s, DN={meta['defaultDN']}, **kw"

    entry = f'''"""Catalog / Plant entry point. Geometry: {part_id}/CUST_{part_id}.py."""
from varmain.custom import *  # type: ignore

import pipe_sizes
from {part_id}.CUST_{part_id} import {py_class}


@activate(  # type: ignore
    Group="{meta['group']}",
    TooltipShort="{meta['short']}",
    TooltipLong="{meta['long']}",
    FirstPortEndtypes="{meta['ends']}",
    LengthUnit="mm",
    Ports="{meta['ports']}",
)
def CUST_{part_id}({activate_sig}):
    preview = bool(kw.get("preview", False))
    return {py_class}({entry_args})
'''

    # Distinct DN (large) sizes present in Excel — drives the insert DN combo so
    # the user can only pick a size that exists (avoids the cubic fallback).
    if reducer:
        sizes = sorted({k[0] for k in table})
        sizes2 = {}
        for dn_l, dn_s in table:
            sizes2.setdefault(dn_l, []).append(dn_s)
        sizes2 = {str(k): sorted(v) for k, v in sizes2.items()}
    else:
        sizes = sorted(table)
        sizes2 = None

    pj = {
        "role": "standard",
        "id": part_id,
        "displayName": meta["display"],
        "category": meta["category"],
        "group": meta["group"],
        "notes": f"NATIVE reconstruction of {meta['shape']}. Source sheet {meta['sheet']}.",
        "catalogParams": ([
            {"name": "DN", "label": "DN", "useSkeletonDN": True, "default": meta["defaultDN"]},
        ] + ([{"name": "DN2", "label": "DN small", "useSkeletonDN2": True, "default": meta["defaultDN2"]}] if reducer else [])),
        "defaultDN": meta["defaultDN"],
        "pressureClass": meta["pclass"],
        "parametricDN": True,
        "primaryEndType": meta["primaryEnd"],
        "pnpClassName": meta["pnp"],
        "shapeName": meta["shape"],
        "excelSheet": meta["sheet"],
        "sizes": sizes,
    }
    if sizes2 is not None:
        pj["sizesByDn"] = sizes2
    if sizes:
        pj["defaultDN"] = sizes[0] if meta["defaultDN"] not in sizes else meta["defaultDN"]

    if dry:
        print(f"  [dry] {part_id}: {len(table)} sizes -> {cls}")
        return
    os.makedirs(geom_dir, exist_ok=True)
    with open(os.path.join(geom_dir, f"CUST_{part_id}.py"), "w", encoding="utf-8", newline="\n") as f:
        f.write(geom)
    with open(os.path.join(part_dir, "catalog_entry.py"), "w", encoding="utf-8", newline="\n") as f:
        f.write(entry)
    with open(os.path.join(part_dir, "part.json"), "w", encoding="utf-8", newline="\n") as f:
        json.dump(pj, f, indent=2)
        f.write("\n")
    print(f"  {part_id}: {len(table)} sizes -> {cls}")


def _meta_from_sheet(name, shape, first_row, ci):
    g = lambda k: (str(first_row[ci[k]]).strip() if k in ci and first_row[ci[k]] else "")
    ends = f"{g('EndType_S1') or g('EndType_S-ALL') or 'FL'},{g('EndType_S2') or g('EndType_S1') or g('EndType_S-ALL') or 'FL'}"
    # Pressure class: single-port sheets use _S1; equal SW/pipe use _S-ALL.
    pc = g("PressureClass_S1") or g("PressureClass_S-ALL") or "150"
    part_id = _sanitize(g("PartFamilyLongDesc") or name).upper()
    # Prefer a compact deterministic id from sheet name.
    part_id = _sanitize(name).upper()
    return {
        "sheet": name, "shape": shape,
        "display": g("ShortDescription") or name.title(),
        "short": (g("ShortDescription") or name)[:40],
        "long": g("PartFamilyLongDesc") or name,
        "category": g("PartCategory") or "Fittings",
        "group": g("PnPClassName") or "Fitting",
        "pnp": g("PnPClassName") or "Fitting",
        "primaryEnd": g("EndType_S1") or "FL",
        "ends": ends, "pclass": pc,
        "ports": g("ConnectionPortCount") or "2",
        "defaultDN": 100, "defaultDN2": 80,
    }, part_id


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=DEFAULT_XLSX)
    ap.add_argument("--dry", action="store_true")
    ap.add_argument("--only", default="", help="comma sheet-name filter")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    only = {s.strip() for s in args.only.split(",") if s.strip()}
    made = 0
    for name in wb.sheetnames:
        if name == "Catalog Data Flag":
            continue
        if only and name not in only:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        ci = _col_index(rows)
        if "ShapeName" not in ci:
            continue
        data = list(_size_rows(rows, ci))
        if not data:
            continue
        shape = str(data[0][ci["ShapeName"]]).strip()
        if shape not in SIMPLE_SHAPES:
            print(f"skip {name}: shape {shape} not in Đợt 1")
            continue
        cls = SIMPLE_SHAPES[shape]
        meta, part_id = _meta_from_sheet(name, shape, data[0], ci)
        reducer = shape in ("CPJRC", "CPJRE")
        table = {}
        for r in data:
            dn, dn2, argt = _row_dims(shape, r, ci)
            key = (dn, dn2) if reducer else dn
            table[key] = [round(x, 4) for x in argt]
        _emit_part(part_id, cls, meta, table, reducer, args.dry)
        made += 1
    print(f"\n{made} part(s) generated from {os.path.basename(args.excel)}")


if __name__ == "__main__":
    main()
